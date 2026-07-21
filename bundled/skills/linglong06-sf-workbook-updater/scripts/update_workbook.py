#!/usr/bin/env python3
"""
SAP SuccessFactors Workbook 更新脚本

功能：
1. 创建 Workbook 副本（带时间戳）
2. 定位目标字段
3. 添加业务规则到 Comments 列
4. 红色高亮新增内容
5. 黄色高亮整行
6. 更新版本历史
"""

import argparse
import shutil
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("请安装 openpyxl: pip install openpyxl")
    exit(1)


# 样式定义
RED_FILL = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
RED_FONT = Font(color='FFFFFF', bold=True, size=10)
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
YELLOW_FONT = Font(color='000000', bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_workbook_copy(source_path: str) -> str:
    """创建带时间戳的 Workbook 副本"""
    source = Path(source_path)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    target = source.parent / f"{source.stem}_updated_{timestamp}{source.suffix}"
    shutil.copy(source, target)
    return str(target)


def find_field_row(sheet, field_id: str) -> int:
    """查找字段所在行（系统字段id 在第3列）"""
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=3).value
        if cell_value and str(cell_value).strip().lower() == field_id.lower():
            return row
    return -1


def find_comments_column(sheet) -> int:
    """查找 Comments 列位置"""
    # 表头通常在第4行
    for row in range(1, min(10, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=row, column=col).value
            if header and 'comment' in str(header).lower():
                return col
    # 默认最后一列
    return sheet.max_column


def update_field_comments(sheet, row: int, comments_col: int, content: str):
    """更新字段 Comments，红色高亮"""
    cell = sheet.cell(row=row, column=comments_col)

    # 追加内容而非覆盖
    existing = cell.value or ""
    if existing:
        new_content = f"{existing}\n\n{content}"
    else:
        new_content = content

    cell.value = new_content
    cell.fill = RED_FILL
    cell.font = RED_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = THIN_BORDER

    # 调整行高
    sheet.row_dimensions[row].height = max(80, sheet.row_dimensions[row].height or 80)


def highlight_row(sheet, row: int, comments_col: int):
    """黄色高亮整行（Comments 列除外）"""
    for col in range(1, comments_col + 1):
        cell = sheet.cell(row=row, column=col)
        if col != comments_col and cell.value:
            cell.fill = YELLOW_FILL
            cell.font = YELLOW_FONT
            cell.border = THIN_BORDER


def update_version_history(wb, version: str, description: str, author: str = "OpenClaw"):
    """更新版本历史"""
    if '版本历史' not in wb.sheetnames:
        return

    sheet = wb['版本历史']
    last_row = sheet.max_row + 1

    # 版本历史列：版本号、版本说明、修改人、邮箱、日期、备注
    data = [version, description, author, f"{author.lower()}@example.com",
            datetime.now().strftime('%Y-%m-%d'), '']

    for col, value in enumerate(data, 2):  # 从B列开始
        cell = sheet.cell(row=last_row, column=col, value=value)
        cell.border = THIN_BORDER
        cell.fill = YELLOW_FILL


def format_business_rule(requirement: dict) -> str:
    """格式化业务规则内容"""
    lines = ["【业务规则-新增】"]

    if 'topic' in requirement:
        lines.append(f"主题: {requirement['topic']}")

    if 'requirement' in requirement:
        lines.append(f"需求: {requirement['requirement']}")

    if 'implementation_note' in requirement:
        lines.append(f"实现方式: {requirement['implementation_note']}")

    if 'trigger_condition' in requirement:
        lines.append(f"触发条件: {requirement['trigger_condition']}")

    lines.append(f"来源会议: {requirement.get('source_meeting', 'N/A')}")
    lines.append(f"更新时间: {datetime.now().strftime('%Y-%m-%d')}")

    return '\n'.join(lines)


def update_workbook(workbook_path: str, requirements: list, source_meeting: str = "") -> str:
    """
    更新 Workbook

    Args:
        workbook_path: Workbook 文件路径
        requirements: 需求列表，每个需求包含：
            - sheet_name: Sheet 名称
            - field_id: 字段 ID
            - requirement: 需求描述
            - implementation_note: 实现建议
            - trigger_condition: 触发条件（可选）
        source_meeting: 来源会议标识

    Returns:
        更新后的文件路径
    """
    # 创建副本
    target_path = create_workbook_copy(workbook_path)
    print(f"创建副本: {target_path}")

    # 加载工作簿
    wb = load_workbook(target_path)

    updated_count = 0

    for req in requirements:
        sheet_name = req.get('sheet_name')
        field_id = req.get('field_id')

        if not sheet_name or not field_id:
            print(f"跳过无效需求: {req}")
            continue

        if sheet_name not in wb.sheetnames:
            print(f"Sheet 不存在: {sheet_name}")
            continue

        sheet = wb[sheet_name]

        # 查找字段行
        row = find_field_row(sheet, field_id)
        if row == -1:
            print(f"字段未找到: {field_id} in {sheet_name}")
            continue

        # 查找 Comments 列
        comments_col = find_comments_column(sheet)

        # 格式化业务规则
        req_with_source = {**req, 'source_meeting': source_meeting}
        business_rule = format_business_rule(req_with_source)

        # 更新 Comments
        update_field_comments(sheet, row, comments_col, business_rule)

        # 高亮整行
        highlight_row(sheet, row, comments_col)

        print(f"已更新: {sheet_name} - {field_id} (行 {row})")
        updated_count += 1

    # 更新版本历史
    update_version_history(wb, "V1.1", f"更新 {updated_count} 个字段配置", "OpenClaw")

    # 保存
    wb.save(target_path)
    print(f"\n✅ Workbook 已更新，共 {updated_count} 个字段")
    print(f"文件位置: {target_path}")

    return target_path


def main():
    parser = argparse.ArgumentParser(description='更新 SAP SuccessFactors Workbook')
    parser.add_argument('workbook', help='Workbook 文件路径')
    parser.add_argument('--requirements', '-r', help='需求 YAML 文件路径')
    parser.add_argument('--sheet', '-s', help='目标 Sheet 名称')
    parser.add_argument('--field', '-f', help='字段 ID')
    parser.add_argument('--requirement', help='需求描述')
    parser.add_argument('--implementation', '-i', help='实现建议')
    parser.add_argument('--source', help='来源会议')

    args = parser.parse_args()

    # 单个需求模式
    if args.sheet and args.field and args.requirement:
        requirements = [{
            'sheet_name': args.sheet,
            'field_id': args.field,
            'requirement': args.requirement,
            'implementation_note': args.implementation or ''
        }]
        update_workbook(args.workbook, requirements, args.source or "")

    # YAML 文件模式
    elif args.requirements:
        import yaml
        with open(args.requirements, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        requirements = data.get('requirements', [])
        update_workbook(args.workbook, requirements, data.get('source_meeting', ''))

    else:
        parser.print_help()


if __name__ == '__main__':
    main()
