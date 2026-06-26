#!/usr/bin/env python3
import os
import sys
import json
import csv
import argparse

# Fallback openpyxl check
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Field definitions for validation and templating
ACTION_FIELDS = {
    "MM": {
        "CREATE_MATERIAL": [
            {"name": "material", "required": False, "description": "Material number (external if supplied)"},
            {"name": "material_type", "required": True, "description": "Material type (e.g. FERT)"},
            {"name": "industry", "required": True, "description": "Industry sector (e.g. 1)"},
            {"name": "description", "required": True, "description": "Material description"},
            {"name": "base_uom", "required": True, "description": "Base unit of measure (e.g. PC)"},
            {"name": "plant", "required": False, "description": "Plant code (e.g. 1000)"},
            {"name": "stor_loc", "required": False, "description": "Storage location (e.g. 0001)"}
        ],
        "CREATE_PO": [
            {"name": "doc_type", "required": True, "description": "PO Doc type (e.g. NB)"},
            {"name": "purch_org", "required": True, "description": "Purchasing Org"},
            {"name": "pur_group", "required": True, "description": "Purchasing Group"},
            {"name": "vendor", "required": True, "description": "Vendor Code"},
            {"name": "material", "required": True, "description": "Material Code"},
            {"name": "plant", "required": True, "description": "Plant"},
            {"name": "quantity", "required": True, "description": "Quantity"},
            {"name": "price", "required": True, "description": "Net Price"}
        ],
        "CHANGE_PO": [
            {"name": "purchase_order", "required": True, "description": "PO number (e.g. 4500001234)"},
            {"name": "item", "required": True, "description": "PO item number"},
            {"name": "quantity", "required": False, "description": "New quantity"},
            {"name": "price", "required": False, "description": "New net price"},
            {"name": "delivery_date", "required": False, "description": "New delivery date (YYYY-MM-DD)"}
        ]
    },
    "SD": {
        "CREATE_ORDER": [
            {"name": "doc_type", "required": True, "description": "Sales doc type (e.g. TA)"},
            {"name": "sales_org", "required": True, "description": "Sales Organization"},
            {"name": "dist_channel", "required": True, "description": "Distribution Channel"},
            {"name": "division", "required": True, "description": "Division"},
            {"name": "customer", "required": True, "description": "Sold-to party customer code"},
            {"name": "material", "required": True, "description": "Material number"},
            {"name": "quantity", "required": True, "description": "Order quantity"},
            {"name": "price", "required": False, "description": "Net price per unit"}
        ],
        "CHANGE_ORDER": [
            {"name": "sales_doc", "required": True, "description": "Sales document number"},
            {"name": "item", "required": True, "description": "Item number"},
            {"name": "quantity", "required": False, "description": "New quantity"},
            {"name": "price", "required": False, "description": "New price"}
        ],
        "CREATE_INVOICE": [
            {"name": "sales_doc", "required": True, "description": "Sales document to bill (REF_DOC in BAPI_BILLINGDOC_CREATEMULTIPLE)"},
            {"name": "billing_type", "required": True, "description": "Billing doc type e.g. F2 (ORDBILLTYP)"},
            {"name": "billing_date", "required": True, "description": "Billing date YYYY-MM-DD (BILL_DATE)"},
            {"name": "ref_doc_ca", "required": False, "description": "Ref doc category J=delivery (REF_DOC_CA)"}
        ],
        "CREATE_DELIVERY": [
            {"name": "sales_doc", "required": True, "description": "Sales order number (SALES_ORDER_ITEMS in BAPI_OUTB_DELIVERY_CREATE_SLS)"},
            {"name": "ship_point", "required": True, "description": "Shipping point (SHIP_POINT)"},
            {"name": "delivery_date", "required": True, "description": "Delivery date YYYY-MM-DD (DUE_DATE)"},
            {"name": "quantity", "required": False, "description": "Delivery quantity (optional, defaults to full order)"}
        ]
    },
    "FI": {
        "POST_DOCUMENT": [
            {"name": "comp_code", "required": True, "description": "Company Code (e.g. 1000)"},
            {"name": "doc_type", "required": True, "description": "Document type (e.g. SA)"},
            {"name": "doc_date", "required": True, "description": "Document date (YYYY-MM-DD)"},
            {"name": "posting_date", "required": True, "description": "Posting date (YYYY-MM-DD)"},
            {"name": "currency", "required": True, "description": "Currency (e.g. BRL)"},
            {"name": "amount", "required": True, "description": "Total amount"},
            {"name": "account", "required": True, "description": "GL account (e.g. 400000)"},
            {"name": "cost_center", "required": False, "description": "Cost center (optional)"}
        ],
        "CHECK_ACCOUNTS": [
            {"name": "account", "required": True, "description": "GL account number (BAPI_GL_GETACCOUNTSALDO)"},
            {"name": "comp_code", "required": True, "description": "Company code"},
            {"name": "fiscal_year", "required": True, "description": "Fiscal year"}
        ],
        "REVERSE_DOCUMENT": [
            {"name": "doc_number", "required": True, "description": "Document number to reverse (OBJ_KEY in BAPI_ACC_DOCUMENT_REV_POST)"},
            {"name": "fiscal_year", "required": True, "description": "Fiscal year"},
            {"name": "comp_code", "required": True, "description": "Company code (COMP_CODE)"},
            {"name": "reason", "required": True, "description": "Reversal reason e.g. 01 (REASON_REV)"},
            {"name": "posting_date", "required": False, "description": "Reversal posting date (PSTNG_DATE)"}
        ]
    },
    "QM": {
        "CREATE_INSPECTION": [
            {"name": "material", "required": True, "description": "Material number (I_MATERIAL in CO_QM_INSPECTION_LOT_CREATE)"},
            {"name": "plant", "required": True, "description": "Plant (I_WERK)"},
            {"name": "insp_type", "required": True, "description": "Inspection type e.g. 01 (I_INSP_TYPE)"},
            {"name": "lot_quantity", "required": True, "description": "Inspection lot quantity"},
            {"name": "base_uom", "required": True, "description": "Base unit of measure"}
        ],
        "RECORD_RESULTS": [
            {"name": "lot_number", "required": True, "description": "Inspection lot number (INSPLOT in BAPI_INSPOPER_RECORDRESULTS)"},
            {"name": "operation", "required": True, "description": "Operation number (INSPOPER)"},
            {"name": "characteristic", "required": True, "description": "MIC characteristic number (INSPCHAR)"},
            {"name": "result_value", "required": True, "description": "Measured value"},
            {"name": "char_attr", "required": False, "description": "Characteristic attribute (CHAR_ATTR)"},
            {"name": "evaluation", "required": False, "description": "Evaluation result code"}
        ]
    },
    "PP": {
        "CREATE_ORDER": [
            {"name": "order_type", "required": True, "description": "Order type (e.g. PP01)"},
            {"name": "plant", "required": True, "description": "Plant"},
            {"name": "material", "required": True, "description": "Material to produce"},
            {"name": "quantity", "required": True, "description": "Total order quantity"},
            {"name": "basic_start", "required": True, "description": "Basic start date (YYYY-MM-DD)"},
            {"name": "basic_end", "required": True, "description": "Basic end date (YYYY-MM-DD)"}
        ],
        "CONFIRM_ORDER": [
            {"name": "order", "required": True, "description": "Production order number (BAPI_PRODORDCONF_CREATE_TT)"},
            {"name": "operation", "required": True, "description": "Operation/activity number"},
            {"name": "confirm_qty", "required": True, "description": "Quantity to confirm"},
            {"name": "yield", "required": False, "description": "Yield quantity"},
            {"name": "scrap", "required": False, "description": "Scrap quantity"},
            {"name": "work_center", "required": False, "description": "Work center (optional)"}
        ],
        "READ_BOM": [
            {"name": "material", "required": True, "description": "Material number (MATNR in CS_BOM_EXPL_MAT_V2)"},
            {"name": "plant", "required": True, "description": "Plant (WERKS)"},
            {"name": "bom_usage", "required": True, "description": "BOM usage e.g. 1 (CAPID)"},
            {"name": "valid_from", "required": False, "description": "Valid-from date YYYY-MM-DD (DATUV)"},
            {"name": "explosion_level", "required": False, "description": "Explosion depth (EMNST)"}
        ],
        "READ_ROUTING": [
            {"name": "material", "required": True, "description": "Material number"},
            {"name": "plant", "required": True, "description": "Plant"},
            {"name": "group", "required": False, "description": "Routing group number"},
            {"name": "group_counter", "required": False, "description": "Routing group counter"}
        ]
    },
    "WM": {
        "GOODS_MOVEMENT": [
            {"name": "material", "required": True, "description": "Material number"},
            {"name": "plant", "required": True, "description": "Plant"},
            {"name": "stor_loc", "required": True, "description": "Storage location"},
            {"name": "move_type", "required": True, "description": "Movement type (e.g. 101)"},
            {"name": "quantity", "required": True, "description": "Quantity"},
            {"name": "base_uom", "required": True, "description": "Base unit of measure"},
            {"name": "move_date", "required": False, "description": "Movement date (YYYY-MM-DD)"}
        ],
        "CREATE_TO": [
            {"name": "warehouse", "required": True, "description": "Warehouse number (I_LGNUM in L_TO_CREATE_SINGLE)"},
            {"name": "move_type", "required": True, "description": "WM movement type e.g. 311 (I_BWLVS)"},
            {"name": "material", "required": True, "description": "Material number (I_MATNR)"},
            {"name": "plant", "required": True, "description": "Plant (I_WERKS)"},
            {"name": "source_stor_type", "required": True, "description": "Source storage type (I_VLTYP)"},
            {"name": "source_stor_bin", "required": True, "description": "Source storage bin (I_VLPLA)"},
            {"name": "dest_stor_type", "required": True, "description": "Destination storage type (I_NLTYP)"},
            {"name": "dest_stor_bin", "required": True, "description": "Destination storage bin (I_NLPLA)"},
            {"name": "quantity", "required": True, "description": "Transfer quantity (I_ANFME)"},
            {"name": "base_uom", "required": True, "description": "Unit of measure (I_ALTME)"}
        ]
    },
    "CO": {
        "CREATE_INTERNAL_ORDER": [
            {"name": "order_type", "required": True, "description": "Order type (e.g. 0100)"},
            {"name": "controlling_area", "required": True, "description": "Controlling area"},
            {"name": "cost_center", "required": True, "description": "Responsible cost center"},
            {"name": "description", "required": True, "description": "Order description"},
            {"name": "currency", "required": True, "description": "Currency (e.g. BRL)"}
        ],
        "ACTIVITY_ALLOC": [
            {"name": "controlling_area", "required": True, "description": "Controlling area (controllingarea in BAPI_CO_ALLOCACTUALS)"},
            {"name": "sender_cost_center", "required": True, "description": "Sender cost center"},
            {"name": "receiver_cost_center", "required": True, "description": "Receiver cost center"},
            {"name": "activity_type", "required": True, "description": "Activity type"},
            {"name": "quantity", "required": True, "description": "Activity quantity"},
            {"name": "period", "required": True, "description": "Fiscal period"},
            {"name": "fiscal_year", "required": True, "description": "Fiscal year"},
            {"name": "version", "required": False, "description": "Plan version (default actual)"}
        ]
    },
    "HCM": {
        "READ_EMPLOYEE": [
            {"name": "employee_id", "required": True, "description": "Personnel number"},
            {"name": "infotype", "required": True, "description": "Infotype (e.g. 0001)"},
            {"name": "subtype", "required": False, "description": "Infotype subtype"},
            {"name": "begin_date", "required": False, "description": "Valid-from date (YYYY-MM-DD)"},
            {"name": "end_date", "required": False, "description": "Valid-to date (YYYY-MM-DD)"}
        ],
        "CREATE_INFOTYPE": [
            {"name": "employee_id", "required": True, "description": "Personnel number"},
            {"name": "infotype", "required": True, "description": "Infotype (e.g. 0002)"},
            {"name": "subtype", "required": False, "description": "Infotype subtype"},
            {"name": "begin_date", "required": True, "description": "Valid-from date (YYYY-MM-DD)"},
            {"name": "end_date", "required": False, "description": "Valid-to date (YYYY-MM-DD)"},
            {"name": "data_json", "required": True, "description": "Infotype data as JSON string"}
        ]
    },
    "BASIS": {
        "CREATE_REQUEST": [
            {"name": "request_type", "required": True, "description": "TR type (e.g. K = Workbench)"},
            {"name": "owner_text", "required": True, "description": "Short description of the request"},
            {"name": "target_system", "required": False, "description": "Target system for delivery"}
        ],
        "RELEASE_REQUEST": [
            {"name": "request", "required": True, "description": "Transport request number"},
            {"name": "task", "required": False, "description": "Release individual task too"}
        ],
        "ST22_SCAN": [
            {"name": "date_from", "required": True, "description": "Start date YYYY-MM-DD (SNAP.DATUM filter)"},
            {"name": "date_to", "required": True, "description": "End date YYYY-MM-DD"},
            {"name": "module", "required": False, "description": "Filter by module e.g. MM (SNAP.PROG)"},
            {"name": "max_dumps", "required": False, "description": "Max dumps to return default 50"}
        ],
        "CODE_ANALYSIS": [
            {"name": "program", "required": True, "description": "ABAP program/class name (TRINT_INSPECT_OBJECTS)"},
            {"name": "check_variant", "required": False, "description": "ATC/SCI variant name"},
            {"name": "object_type", "required": False, "description": "Object type e.g. CLAS, PROG"}
        ],
        "CODE_SEARCH": [
            {"name": "query", "required": True, "description": "Search term / regex pattern"},
            {"name": "mode", "required": False, "description": "Search mode: STRING, REGEX, or PCRE (default STRING)"},
            {"name": "object_type", "required": False, "description": "Object type filter: CLAS,INTF,PROG,TYPE,FUGR,DDLS,DDLX,DCLS,BDEV,XSLT,STRU,DTAB (blank=all)"},
            {"name": "package", "required": False, "description": "Limit to package (blank=all)"},
            {"name": "owner", "required": False, "description": "Limit to owner username (blank=all)"},
            {"name": "max_results", "required": False, "description": "Max hits (default 200)"},
            {"name": "ignore_case", "required": False, "description": "Case-insensitive search (X=yes, default true)"},
            {"name": "parallel", "required": False, "description": "Enable parallel processing (X=yes)"},
            {"name": "from_date", "required": False, "description": "Changed after date (YYYY-MM-DD)"},
            {"name": "to_date", "required": False, "description": "Changed before date (YYYY-MM-DD)"}
        ],
        "CODE_SEARCH_STATS": [
            {"name": "package", "required": False, "description": "Filter by package (blank=all)"},
            {"name": "owner", "required": False, "description": "Filter by owner username (blank=all)"}
        ],
        "CODE_SEARCH_ADT": [
            {"name": "query", "required": True, "description": "Search term / regex pattern"},
            {"name": "mode", "required": False, "description": "Search mode: STRING, REGEX, or PCRE"},
            {"name": "object_type", "required": False, "description": "Object type filter (blank=all 12 types)"},
            {"name": "max_results", "required": False, "description": "Max hits (default 200)"}
        ]
    }
}

class XlsToBapi:
    def __init__(self, module, action_name):
        self.module = module.upper()
        self.action_name = action_name.upper()
        self.fields = ACTION_FIELDS.get(self.module, {}).get(self.action_name, [])

    def get_template_fields(self):
        return [f["name"] for f in self.fields]

    def validate_row(self, row, row_num):
        errors = []
        for field in self.fields:
            name = field["name"]
            val = row.get(name, "")
            if field["required"] and not val:
                errors.append(f"Row {row_num}: Missing required field '{name}'")
        return errors

    def parse_csv(self, filepath):
        results = []
        errors = []
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            # Normalize headers
            headers = [h.strip().lower() for h in reader.fieldnames]
            reader.fieldnames = headers
            
            for idx, row in enumerate(reader, start=2):
                # Clean row values
                cleaned_row = {k.strip().lower(): v.strip() for k, v in row.items() if k}
                # Map to official case-sensitive field names
                mapped_row = {}
                for field in self.fields:
                    fname = field["name"]
                    mapped_row[fname] = cleaned_row.get(fname.lower(), "")
                
                row_errors = self.validate_row(mapped_row, idx)
                if row_errors:
                    errors.extend(row_errors)
                else:
                    results.append(mapped_row)
        return results, errors

    def parse_excel(self, filepath):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is not installed. Please use CSV or install openpyxl.")

        results = []
        errors = []
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val).strip().lower() if val else "")

        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                if header:
                    val = ws.cell(row=row_idx, column=col_idx).value
                    row_data[header] = str(val).strip() if val is not None else ""
            
            mapped_row = {}
            for field in self.fields:
                fname = field["name"]
                mapped_row[fname] = row_data.get(fname.lower(), "")

            row_errors = self.validate_row(mapped_row, row_idx)
            if row_errors:
                errors.extend(row_errors)
            else:
                results.append(mapped_row)
                
        return results, errors

    def generate_csv_template(self, output_path):
        headers = self.get_template_fields()
        with open(output_path, mode='w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            # Write descriptions as second line comments/samples
            descriptions = [f["description"] for f in self.fields]
            writer.writerow(descriptions)

    def generate_excel_template(self, output_path):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is not installed. Please use CSV template generator.")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Template_{self.action_name}"
        
        for idx, field in enumerate(self.fields, start=1):
            cell_header = ws.cell(row=1, column=idx, value=field["name"])
            cell_header.font = openpyxl.styles.Font(bold=True)
            ws.cell(row=2, column=idx, value=field["description"])
            
        wb.save(output_path)

def main():
    parser = argparse.ArgumentParser(description="XLS/CSV Converter for BAPI Payloads")
    parser.add_argument("command", choices=["convert", "template", "validate"])
    parser.add_argument("--input", help="Path to input XLS/CSV file")
    parser.add_argument("--output", help="Path to output JSON/XLS/CSV file")
    parser.add_argument("--module", required=True, help="SAP Module (e.g. MM)")
    parser.add_argument("--action", required=True, help="Action name (e.g. CREATE_MATERIAL)")
    
    args = parser.parse_args()
    
    # Check action existence
    if args.module.upper() not in ACTION_FIELDS or args.action.upper() not in ACTION_FIELDS[args.module.upper()]:
        available = []
        for mod, acts in ACTION_FIELDS.items():
            for act in acts:
                available.append(f"{mod}_{act}")
        print(f"Error: Module/Action not supported. Available: {', '.join(available)}", file=sys.stderr)
        sys.exit(1)
        
    converter = XlsToBapi(args.module, args.action)
    
    if args.command == "template":
        if not args.output:
            print("Error: --output is required for template command", file=sys.stderr)
            sys.exit(1)
        ext = os.path.splitext(args.output)[1].lower()
        if ext == '.csv':
            converter.generate_csv_template(args.output)
            print(f"CSV Template written to {args.output}")
        elif ext in ['.xlsx', '.xls']:
            try:
                converter.generate_excel_template(args.output)
                print(f"Excel Template written to {args.output}")
            except ImportError as e:
                print(e, file=sys.stderr)
                sys.exit(1)
        else:
            print("Error: Template extension must be .csv or .xlsx", file=sys.stderr)
            sys.exit(1)
            
    elif args.command in ["convert", "validate"]:
        if not args.input:
            print("Error: --input is required for convert/validate command", file=sys.stderr)
            sys.exit(1)
            
        ext = os.path.splitext(args.input)[1].lower()
        results = []
        errors = []
        
        try:
            if ext == '.csv':
                results, errors = converter.parse_csv(args.input)
            elif ext in ['.xlsx', '.xls']:
                results, errors = converter.parse_excel(args.input)
            else:
                print("Error: Input file must be .csv or .xlsx", file=sys.stderr)
                sys.exit(1)
        except Exception as e:
            print(f"Error reading file: {e}", file=sys.stderr)
            sys.exit(1)
            
        if errors:
            print("Validation errors found:")
            for err in errors:
                print(f" - {err}")
            sys.exit(1)
            
        print(f"Validation successful. Processed {len(results)} rows.")
        
        if args.command == "convert":
            if not args.output:
                # Output to stdout
                print(json.dumps(results, indent=2))
            else:
                with open(args.output, 'w', encoding='utf-8') as f:
                    json.dump(results, f, indent=2)
                print(f"Converted payload saved to {args.output}")

if __name__ == "__main__":
    main()
